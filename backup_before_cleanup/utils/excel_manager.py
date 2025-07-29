import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from logs.logger import log
from utils.db_manager import get_all_trades, get_statistics_by_symbol

EXCEL_DIR = "reports/excel_reports"

def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def export_trades_to_excel(filename=None):
    """Eksportuje SVE trejdove iz baze u jedan Excel fajl."""
    ensure_dir(EXCEL_DIR)
    if not filename:
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"trades_{today}.xlsx"
    path = os.path.join(EXCEL_DIR, filename)

    trades = get_all_trades()
    if not trades:
        log.warning("⚠️ Nema trejdova za eksport u Excel.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    # Header
    headers = ["ID", "Symbol", "Entry Time", "Exit Time", "Type", "Entry Price", "Exit Price", "SL", "TP", "Profit", "Duration", "Opened By", "Outcome"]
    ws.append(headers)

    for t in trades:
        ws.append(list(t))

    wb.save(path)
    log.info(f"✅ Excel izveštaj sačuvan: {path}")
    return path

def export_statistics_to_excel(filename=None):
    """Eksportuje sumarnu statistiku po simbolima u Excel."""
    ensure_dir(EXCEL_DIR)
    if not filename:
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"statistics_{today}.xlsx"
    path = os.path.join(EXCEL_DIR, filename)

    stats = get_statistics_by_symbol()
    if not stats:
        log.warning("⚠️ Nema statistike za eksport u Excel.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"

    # Header
    headers = ["Symbol", "Total Trades", "Total Profit", "Avg Profit", "Max Profit", "Min Profit"]
    ws.append(headers)

    for s in stats:
        ws.append([
            s["symbol"],
            s["total_trades"],
            s["total_profit"],
            s["avg_profit"],
            s["max_profit"],
            s["min_profit"]
        ])

    wb.save(path)
    log.info(f"✅ Excel statistika sačuvana: {path}")
    return path

def read_excel(filepath):
    """Učitaj i vrati sadržaj Excel fajla kao listu redova."""
    if not os.path.exists(filepath):
        log.error(f"❌ Fajl ne postoji: {filepath}")
        return []
    wb = load_workbook(filepath)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
    log.info(f"📖 Pročitano {len(rows)} redova iz {filepath}")
    return rows

if __name__ == "__main__":
    # Test eksport funkcija
    export_trades_to_excel()
    export_statistics_to_excel()
