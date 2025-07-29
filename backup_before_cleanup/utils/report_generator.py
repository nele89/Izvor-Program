import openpyxl
from openpyxl.styles import Font
import os
from datetime import datetime
from logs.logger import log

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


def save_report(data, report_type="daily", summary=False, generate_pdf=True):
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")

    # Mape foldera prema tipu izveštaja
    report_folders = {
        "daily": "reports/daily_reports",
        "weekly": "reports/weekly_reports",
        "monthly": "reports/monthly_reports",
        "yearly": "reports/yearly_reports"
    }

    folder = report_folders.get(report_type.lower(), report_folders["daily"])
    os.makedirs(folder, exist_ok=True)

    filename_base = f"{report_type}_report_{timestamp}"
    excel_path = os.path.join(folder, f"{filename_base}.xlsx")

    # Excel generisanje
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Report"

        bold_font = Font(bold=True)
        sheet.append(["Parametar", "Vrednost"])
        sheet["A1"].font = bold_font
        sheet["B1"].font = bold_font

        for key, value in data.items():
            sheet.append([key, value])

        if summary:
            sheet.append([])
            sheet.append(["Ukupno stavki", len(data)])

        wb.save(excel_path)
        log.info(f"✅ Excel izveštaj sačuvan: {excel_path}")
    except Exception as e:
        log.error(f"❌ Greška pri snimanju Excel izveštaja: {e}")

    # PDF generisanje
    if generate_pdf and PDF_AVAILABLE:
        try:
            pdf_path = os.path.join(folder, f"{filename_base}.pdf")
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt=f"{report_type.capitalize()} izveštaj", ln=True, align='C')
            pdf.ln(5)

            for key, value in data.items():
                pdf.cell(200, 8, txt=f"{key}: {value}", ln=True)

            if summary:
                pdf.ln(5)
                pdf.cell(200, 8, txt=f"Ukupno stavki: {len(data)}", ln=True)

            pdf.output(pdf_path)
            log.info(f"📄 PDF izveštaj sačuvan: {pdf_path}")
        except Exception as e:
            log.error(f"❌ Greška pri snimanju PDF izveštaja: {e}")

    return excel_path
